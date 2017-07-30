# coding=utf-8
from django.contrib.auth.models import User
from django.test import TestCase

from events.models import Survey


class ExportSurveyResultsAsExcelTests(TestCase):

    def setUp(self):
        self.user1 = User.objects.create_user(
            username='testuser1',
            email='test1@example.com',
            password='abcd1234',
        )
        self.user1.is_staff = True
        self.user1.save()

        self.client.login(
            username='testuser1',
            password='abcd1234',
        )

        self.user2 = User.objects.create_user(
            username='testuser2',
            email='test2@example.com',
            password='abcd1234',
        )

        self.user3 = User.objects.create_user(
            username='testuser3',
            email='test3@example.com',
            password='abcd1234',
        )

        self.survey = Survey.objects.create(name="Feedback")
        question1 = self.survey.survey_questions.create(
            category='O',
            text=u"ما هي هوايتك المفضلة؟",
        )
        question2 = self.survey.survey_questions.create(
            category='C',
            choices=u"لندن\nباريس",
            text=u"إلى أين تود الذهاب في الإجازة القادمة؟",
        )

        response1 = self.survey.responses.create(
            user=self.user2,
        )
        response1.answers.create(
            question=question1,
            text_value=u"الرسم",
        )
        response1.answers.create(
            question=question2,
            text_value=u"لندن",
        )

        response2 = self.survey.responses.create(
            user=self.user3,
        )
        response2.answers.create(
            question=question1,
            text_value=u"السباحة",
        )
        response2.answers.create(
            question=question2,
            text_value=u"باريس",
        )

    def test_view_returns_appropriate_response(self):
        response = self.client.get("/admin/events/survey/{}/results/".format(self.survey.id))

        self.assertEqual(response.status_code, 200)

        self.assertEqual(
            response.get('Content-Disposition'),
            "attachment; filename={}".format("Survey Results.xlsx")
        )

        self.assertEqual(
            response.get('Content-Type'),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8"
        )

    def test_excel_file_contains_proper_info(self):
        response = self.client.get("/admin/events/survey/{}/results/".format(self.survey.id))

        from StringIO import StringIO
        import openpyxl
        wb = openpyxl.load_workbook(StringIO(response.content))
        ws = wb.active

        self.assertEqual(
            ws["A1:A4"],
            [u"التاريخ و الوقت", u"المستخدم", u"ما هي هوايتك المفضلة؟", u"إلى أين تود الذهاب في الإجازة القادمة؟"],
        )

        self.assertEqual(
            ws["B2:B4"],
            ["testuser2", u"الرسم", u"لندن"],
        )

        self.assertEqual(
            ws["C2:C4"],
            ["testuser3", u"السباحة", u"باريس"],
        )
